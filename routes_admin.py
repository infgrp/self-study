"""
관리자 Blueprint - 교사 계정 승인, 비밀번호 변경, 전체 현황, 사용자 관리
"""

import os
import secrets
import sqlite3
import tempfile
from datetime import datetime
from io import BytesIO
from flask import Blueprint, render_template, redirect, url_for, request, flash, send_file
from flask_login import login_required, current_user
from models import db, User, Schedule, Attendance, StudyLog, StudentRoom, StudyRoom, StudyApplication, AttendanceLog, Holiday, StudyPeriodSetting, SystemSetting
import settings
from settings import SETTINGS_SCHEMA
from validators import validate_password, validate_student_id, generate_temp_password
from time_utils import validate_time_str, parse_time_str
from audit import log_audit
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# SQLite DB 파일 경로
_HERE   = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(_HERE, 'instance', 'self_study.db')

admin_bp = Blueprint('admin_bp', __name__)


@admin_bp.before_request
@login_required
def check_admin():
    if current_user.role != 'admin':
        flash('관리자만 접근할 수 있습니다.', 'danger')
        return redirect(url_for('auth.login'))


@admin_bp.route('/')
def dashboard():
    total_students = User.query.filter_by(role='student').count()
    total_teachers = User.query.filter_by(role='teacher').count()
    pending_count  = User.query.filter_by(role='teacher', is_approved=False).count()

    return render_template('admin/dashboard.html',
                           total_students=total_students,
                           total_teachers=total_teachers,
                           pending_count=pending_count)


@admin_bp.route('/teachers')
def teachers():
    """교사 목록 및 승인 관리"""
    status_filter = request.args.get('status', 'all')  # all / pending / approved

    query = User.query.filter_by(role='teacher')
    if status_filter == 'pending':
        query = query.filter_by(is_approved=False)
    elif status_filter == 'approved':
        query = query.filter_by(is_approved=True)

    teacher_list = query.order_by(User.is_approved, User.id).all()
    pending_count = User.query.filter_by(role='teacher', is_approved=False).count()

    return render_template('admin/teachers.html',
                           teachers=teacher_list,
                           status_filter=status_filter,
                           pending_count=pending_count)


@admin_bp.route('/teachers/<int:user_id>/approve', methods=['POST'])
def approve_teacher(user_id):
    """교사 계정 승인"""
    teacher = db.session.get(User, user_id)
    if not teacher or teacher.role != 'teacher':
        flash('존재하지 않는 교사 계정입니다.', 'danger')
        return redirect(url_for('admin_bp.teachers'))

    teacher.is_approved = True
    db.session.commit()
    flash(f'"{teacher.name}" 교사 계정이 승인되었습니다.', 'success')
    return redirect(url_for('admin_bp.teachers'))


@admin_bp.route('/teachers/<int:user_id>/reject', methods=['POST'])
def reject_teacher(user_id):
    """교사 계정 삭제 (승인 거부)"""
    teacher = db.session.get(User, user_id)
    if not teacher or teacher.role != 'teacher':
        flash('존재하지 않는 교사 계정입니다.', 'danger')
        return redirect(url_for('admin_bp.teachers'))

    name = teacher.name
    # FK: AttendanceLog.changed_by 참조 정리 후 삭제
    AttendanceLog.query.filter_by(changed_by=user_id).update(
        {'changed_by': None}, synchronize_session=False
    )
    db.session.delete(teacher)
    db.session.commit()
    flash(f'"{name}" 교사 계정이 삭제(거부)되었습니다.', 'warning')
    return redirect(url_for('admin_bp.teachers'))


@admin_bp.route('/teachers/<int:user_id>/revoke', methods=['POST'])
def revoke_teacher(user_id):
    """이미 승인된 교사의 권한 취소"""
    teacher = db.session.get(User, user_id)
    if not teacher or teacher.role != 'teacher':
        flash('존재하지 않는 교사 계정입니다.', 'danger')
        return redirect(url_for('admin_bp.teachers'))

    teacher.is_approved = False
    db.session.commit()
    flash(f'"{teacher.name}" 교사의 접근 권한이 취소되었습니다.', 'warning')
    return redirect(url_for('admin_bp.teachers'))


@admin_bp.route('/system-settings', methods=['GET', 'POST'])
def system_settings():
    """시스템 운영 설정 - 정책값 일괄 조회/수정"""
    if request.method == 'POST':
        errors = []
        changed = 0
        for spec in SETTINGS_SCHEMA:
            key = spec['key']
            raw = request.form.get(key)
            if raw is None:
                continue
            raw = raw.strip()

            # 타입별 검증
            if spec['type'] == 'int':
                try:
                    val = int(raw)
                except ValueError:
                    errors.append(f"{spec['description']}: 정수가 아닙니다 ({raw!r})")
                    continue
                if spec['min'] is not None and val < spec['min']:
                    errors.append(f"{spec['description']}: 최소 {spec['min']} 이상이어야 합니다")
                    continue
                if spec['max'] is not None and val > spec['max']:
                    errors.append(f"{spec['description']}: 최대 {spec['max']} 이하여야 합니다")
                    continue
                normalized = str(val)
            elif spec['type'] == 'bool':
                # 체크박스 미체크 시 form에서 키 자체가 누락되므로 별도 처리
                normalized = 'true' if raw.lower() in ('true', '1', 'on', 'yes') else 'false'
            else:
                normalized = raw

            row = SystemSetting.query.filter_by(key=key).first()
            if row and row.value != normalized:
                row.value = normalized
                row.updated_by = current_user.id
                changed += 1

        # 체크박스가 unchecked로 들어온 bool 키 처리
        for spec in SETTINGS_SCHEMA:
            if spec['type'] != 'bool':
                continue
            if request.form.get(spec['key']) is None:
                row = SystemSetting.query.filter_by(key=spec['key']).first()
                if row and row.value != 'false':
                    row.value = 'false'
                    row.updated_by = current_user.id
                    changed += 1

        if errors:
            db.session.rollback()
            for msg in errors:
                flash(msg, 'danger')
        else:
            db.session.commit()
            if changed:
                flash(f'시스템 설정 {changed}개 항목이 변경되었습니다.', 'success')
            else:
                flash('변경된 항목이 없습니다.', 'info')
        return redirect(url_for('admin_bp.system_settings'))

    rows = []
    for spec in SETTINGS_SCHEMA:
        row = SystemSetting.query.filter_by(key=spec['key']).first()
        rows.append({
            'key': spec['key'],
            'description': spec['description'],
            'type': spec['type'],
            'min': spec['min'],
            'max': spec['max'],
            'value': row.value if row else spec['default'],
            'updated_at': row.updated_at if row else None,
            'updated_by': row.updater.name if (row and row.updater) else None,
        })
    return render_template('admin/system_settings.html', rows=rows)


@admin_bp.route('/change-password', methods=['GET', 'POST'])
def change_password():
    """관리자 비밀번호 변경"""
    if request.method == 'POST':
        current_pw  = request.form.get('current_password', '')
        new_pw      = request.form.get('new_password', '')
        confirm_pw  = request.form.get('confirm_password', '')

        if not current_user.check_password(current_pw):
            flash('현재 비밀번호가 올바르지 않습니다.', 'danger')
            return render_template('admin/change_password.html')

        if new_pw != confirm_pw:
            flash('새 비밀번호가 일치하지 않습니다.', 'danger')
            return render_template('admin/change_password.html')

        ok, err = validate_password(new_pw)
        if not ok:
            flash(err, 'danger')
            return render_template('admin/change_password.html')

        current_user.set_password(new_pw)
        db.session.commit()
        flash('비밀번호가 변경되었습니다. 다시 로그인해 주세요.', 'success')
        return redirect(url_for('admin_bp.dashboard'))

    return render_template('admin/change_password.html')


# ── 사용자 관리 (학생·교사 삭제 / 비밀번호 초기화) ─────────────

def _temp_password():
    """영문+숫자 혼합 임시 비밀번호 생성 (길이는 SystemSetting에서 조회)"""
    return generate_temp_password()


@admin_bp.route('/users')
def users():
    """학생·교사 전체 목록"""
    role_filter = request.args.get('role', 'all')
    search      = request.args.get('q', '').strip()

    query = User.query.filter(User.role != 'admin')
    if role_filter == 'student':
        query = query.filter_by(role='student')
    elif role_filter == 'teacher':
        query = query.filter_by(role='teacher')
    if search:
        query = query.filter(
            (User.name.contains(search)) | (User.username.contains(search))
        )

    user_list = query.order_by(User.role, User.name).all()
    return render_template('admin/users.html',
                           user_list=user_list,
                           role_filter=role_filter,
                           search=search)


@admin_bp.route('/users/<int:user_id>/delete', methods=['POST'])
def delete_user(user_id):
    """특정 학생·교사 계정 및 관련 데이터 전체 삭제"""
    user = db.session.get(User, user_id)
    if not user or user.role == 'admin':
        flash('삭제할 수 없는 계정입니다.', 'danger')
        return redirect(url_for('admin_bp.users'))

    # 연관 데이터 먼저 삭제 (FK 제약 방지)
    StudyApplication.query.filter_by(user_id=user_id).delete()
    StudentRoom.query.filter_by(user_id=user_id).delete()
    # AttendanceLog는 ORM cascade가 bulk delete를 우회하므로 명시적으로 삭제
    att_ids = [a.id for a in Attendance.query.filter_by(user_id=user_id)
                                             .with_entities(Attendance.id).all()]
    if att_ids:
        AttendanceLog.query.filter(
            AttendanceLog.attendance_id.in_(att_ids)
        ).delete(synchronize_session=False)
    Attendance.query.filter_by(user_id=user_id).delete()
    # 교사 삭제 시 AttendanceLog.changed_by 참조 정리
    if user.role == 'teacher':
        AttendanceLog.query.filter_by(changed_by=user_id).update(
            {'changed_by': None}, synchronize_session=False
        )
    StudyLog.query.filter_by(user_id=user_id).delete()
    Schedule.query.filter_by(user_id=user_id).delete()

    name = user.name
    target_username = user.username
    target_role = user.role
    db.session.delete(user)
    db.session.commit()
    log_audit('admin.account_delete', level='warning',
              admin=current_user.username, target=target_username,
              target_role=target_role)
    flash(f'"{name}" 계정과 모든 관련 데이터가 삭제되었습니다.', 'warning')
    return redirect(url_for('admin_bp.users'))


@admin_bp.route('/new-year', methods=['GET', 'POST'])
def new_year():
    """새 학년도 초기화 — 학생 데이터 전체 삭제, 교사/시설 설정 보존"""
    CONFIRM_PHRASE = '새학년도초기화'

    # ── 현황 집계 (GET/POST 공통) ──
    stats = {
        'students':     User.query.filter_by(role='student').count(),
        'applications': StudyApplication.query.count(),
        'attendances':  Attendance.query.count(),
        'study_logs':   StudyLog.query.count(),
        'att_logs':     AttendanceLog.query.count(),
    }

    if request.method == 'POST':
        phrase = request.form.get('confirm_phrase', '').strip()
        if phrase != CONFIRM_PHRASE:
            flash(f'확인 문구가 올바르지 않습니다. 정확히 "{CONFIRM_PHRASE}"를 입력하세요.', 'danger')
            return render_template('admin/new_year.html',
                                   stats=stats,
                                   confirm_phrase=CONFIRM_PHRASE,
                                   year=datetime.now().year)

        # ── 순서 중요: FK 참조 순으로 삭제 ──
        AttendanceLog.query.delete()
        Attendance.query.delete()
        StudyApplication.query.delete()
        StudyLog.query.delete()
        Schedule.query.filter(
            Schedule.user_id.in_(
                db.session.query(User.id).filter_by(role='student')
            )
        ).delete(synchronize_session='fetch')
        StudentRoom.query.delete()

        # 학생 계정 삭제
        User.query.filter_by(role='student').delete()
        db.session.commit()
        log_audit('admin.new_year_reset', level='warning',
                  admin=current_user.username,
                  students=stats['students'], attendances=stats['attendances'],
                  applications=stats['applications'], study_logs=stats['study_logs'])

        flash(
            f'새 학년도 초기화 완료 — '
            f'학생 {stats["students"]}명, 출결 {stats["attendances"]}건, '
            f'신청 {stats["applications"]}건, 학습기록 {stats["study_logs"]}건 삭제됨.',
            'success'
        )
        return redirect(url_for('admin_bp.dashboard'))

    return render_template('admin/new_year.html',
                           stats=stats,
                           confirm_phrase=CONFIRM_PHRASE,
                           year=datetime.now().year)


@admin_bp.route('/new-year/backup')
def new_year_backup():
    """새 학년도 초기화 전 전체 데이터 Excel 백업"""
    wb = openpyxl.Workbook()

    # ── 공통 스타일 ──
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    hdr_align = Alignment(horizontal='center', vertical='center')

    def make_header(ws, cols):
        ws.append(cols)
        for cell in ws[1]:
            cell.fill  = hdr_fill
            cell.font  = hdr_font
            cell.alignment = hdr_align
        ws.freeze_panes = 'A2'

    # ── Sheet 1: 학생 명단 ──
    # 보안: 비밀번호 해시는 백업에 포함하지 않음. 복원 시 임시 비번이 새로 발급된다.
    ws1 = wb.active
    ws1.title = '학생명단'
    make_header(ws1, ['학번', '이름', '학년', '반', '성별', '아이디'])
    for s in User.query.filter_by(role='student').order_by(
            User.grade, User.class_num, User.student_id).all():
        ws1.append([
            s.student_id,
            s.name,
            s.grade,
            s.class_num,
            '남' if s.gender == 'M' else '여',
            s.username,
        ])

    # ── Sheet 2: 출결 요약 (학생 × 월별) ──
    ws2 = wb.create_sheet('출결요약')
    from collections import defaultdict
    import calendar as cal_mod
    summary = defaultdict(lambda: {'present': 0, 'late': 0, 'absent': 0,
                                   'early_leave': 0, 'approved_leave': 0,
                                   'after_school': 0})
    for att in Attendance.query.all():
        key = (att.user_id, att.date.year, att.date.month)
        if att.status in summary[key]:
            summary[key][att.status] += 1

    make_header(ws2, ['학번', '이름', '학년', '반', '연도', '월',
                      '출석', '지각', '결석', '조퇴', '출석인정', '방과후출결',
                      '신청', '참여율(%)'])
    students = {s.id: s for s in User.query.filter_by(role='student').all()}
    app_count = defaultdict(int)
    for app in StudyApplication.query.all():
        key = (app.user_id, app.date.year, app.date.month)
        app_count[key] += 1

    for (uid, yr, mo), cnt in sorted(summary.items()):
        s = students.get(uid)
        if not s:
            continue
        applied      = app_count[(uid, yr, mo)]
        present      = cnt['present']
        late         = cnt['late']
        approved     = cnt['approved_leave']
        after_school = cnt['after_school']
        # 참여율: 출석 + 지각 + 출석인정 + 방과후출결인정 (다른 집계와 정합)
        rate = round((present + late + approved + after_school) / applied * 100) if applied else 0
        ws2.append([
            s.student_id, s.name, s.grade, s.class_num,
            yr, mo, present, late, cnt['absent'],
            cnt['early_leave'], approved, after_school, applied, rate
        ])

    # ── Sheet 3: 학습 기록 ──
    ws3 = wb.create_sheet('학습기록')
    make_header(ws3, ['학번', '이름', '학년', '반', '날짜', '과목', '학습시간(분)', '메모'])
    for log in StudyLog.query.order_by(StudyLog.date.desc()).all():
        s = students.get(log.user_id)
        if not s:
            continue
        ws3.append([
            s.student_id, s.name, s.grade, s.class_num,
            log.date.isoformat(), log.subject, log.duration, log.memo or ''
        ])

    # ── Sheet 4: 자습 신청 현황 ──
    ws4 = wb.create_sheet('자습신청')
    make_header(ws4, ['학번', '이름', '학년', '반', '날짜', '교시'])
    for app in StudyApplication.query.order_by(
            StudyApplication.date.desc(), StudyApplication.period).all():
        s = students.get(app.user_id)
        if not s:
            continue
        ws4.append([
            s.student_id, s.name, s.grade, s.class_num,
            app.date.isoformat(), app.period
        ])

    # ── Sheet 5: 출결 상세 (복원용 원시 데이터) ──
    ws5 = wb.create_sheet('출결상세')
    make_header(ws5, ['학번', '이름', '날짜', '교시', '상태', '출석시각', '퇴실시각', '조퇴사유', '자습공간명', '자습시간(분)'])
    room_map = {r.id: r.name for r in StudyRoom.query.all()}
    STATUS_KO = {
        'present':        '출석',
        'late':           '지각',
        'absent':         '결석',
        'early_leave':    '조퇴',
        'approved_leave': '출석인정',
        'after_school':   '방과후출결인정',
    }
    for att in Attendance.query.order_by(Attendance.date, Attendance.period).all():
        s = students.get(att.user_id)
        if not s:
            continue
        ws5.append([
            s.student_id,
            s.name,
            att.date.isoformat(),
            att.period,
            STATUS_KO.get(att.status, att.status),
            att.checked_at.strftime('%Y-%m-%d %H:%M:%S')     if att.checked_at     else '',
            att.checked_out_at.strftime('%Y-%m-%d %H:%M:%S') if att.checked_out_at else '',
            att.early_leave_note or '',
            room_map.get(att.study_room_id, '') if att.study_room_id else '',
            att.study_minutes if att.study_minutes is not None else '',
        ])

    # ── Sheet 6: 교사 명단 (아이디·이름·승인여부·담당학년) ──
    # 보안: 비밀번호 해시는 백업에 포함하지 않음.
    ws6 = wb.create_sheet('교사명단')
    make_header(ws6, ['아이디', '이름', '승인여부', '담당학년'])
    for t in User.query.filter_by(role='teacher').order_by(User.name).all():
        ws6.append([
            t.username,
            t.name,
            '승인' if t.is_approved else '미승인',
            t.assigned_grade if t.assigned_grade is not None else '',
        ])

    # ── Sheet 7: 관리자 계정 (아이디·이름) ──
    # 보안: 비밀번호 해시는 백업에 포함하지 않음.
    ws7 = wb.create_sheet('관리자')
    make_header(ws7, ['아이디', '이름'])
    for a in User.query.filter_by(role='admin').order_by(User.id).all():
        ws7.append([a.username, a.name])

    # ── Sheet 8: 자습 시간 설정 ──
    ws8 = wb.create_sheet('자습시간설정')
    make_header(ws8, ['요일구분', '교시', '시작시각', '종료시각', '활성화'])
    for sp in StudyPeriodSetting.query.order_by(
            StudyPeriodSetting.day_type, StudyPeriodSetting.period).all():
        ws8.append([
            sp.day_type,
            sp.period,
            sp.start_time,
            sp.end_time,
            'Y' if sp.is_active else 'N',
        ])

    # ── Sheet 9: 공휴일 목록 ──
    ws9 = wb.create_sheet('공휴일')
    make_header(ws9, ['날짜', '공휴일명'])
    for h in Holiday.query.order_by(Holiday.date).all():
        ws9.append([h.date.isoformat(), h.name])

    # ── Sheet 10: 자습실 목록 ──
    # 보안: QR 토큰은 백업에 포함하지 않음. 복원 시 새 토큰이 자동 생성되며, QR 재인쇄가 필요하다.
    ws10 = wb.create_sheet('자습실목록')
    make_header(ws10, ['자습실명', '전체정원', '남학생정원', '여학생정원', '활성화', '순서'])
    all_rooms = StudyRoom.query.order_by(StudyRoom.order).all()
    for r in all_rooms:
        ws10.append([
            r.name,
            r.capacity,
            r.male_capacity,
            r.female_capacity,
            'Y' if r.is_active else 'N',
            r.order,
        ])

    # ── Sheet 11: 자습실 배정 및 좌석 위치 ──
    ws11 = wb.create_sheet('자습실배정')
    make_header(ws11, ['학번', '이름', '자습실명', '좌석번호', '위치X(%)', '위치Y(%)'])
    room_map = {r.id: r.name for r in all_rooms}
    for sr in StudentRoom.query.all():
        s = students.get(sr.user_id)
        if not s:
            continue
        ws11.append([
            s.student_id,
            s.name,
            room_map.get(sr.study_room_id, ''),
            sr.seat_number if sr.seat_number is not None else '',
            round(sr.pos_x, 4) if sr.pos_x is not None else '',
            round(sr.pos_y, 4) if sr.pos_y is not None else '',
        ])

    # ── 12. 방과후 수업 스케줄 ──
    DAY_NAMES = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금'}
    ws12 = wb.create_sheet('방과후수업')
    make_header(ws12, ['학번', '이름', '학년', '반', '요일', '요일번호(0=월)', '교시'])
    for sc in Schedule.query.order_by(Schedule.user_id, Schedule.day_of_week, Schedule.period).all():
        s = students.get(sc.user_id)
        if not s:
            continue
        ws12.append([
            s.student_id,
            s.name,
            s.grade,
            s.class_num,
            DAY_NAMES.get(sc.day_of_week, str(sc.day_of_week)),
            sc.day_of_week,
            sc.period,
        ])

    # 열 너비 자동 조정
    for ws in [ws1, ws2, ws3, ws4, ws5, ws6, ws7, ws8, ws9, ws10, ws11, ws12]:
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    # 파일 전송
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    now_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'자율학습_백업_{now_str}.xlsx'
    return send_file(buf, as_attachment=True,
                     download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Excel 복원 헬퍼들 ─────────────────────────────────────────
# restore()가 너무 비대해져서 시트별로 분리. 각 헬퍼는 (wb, result, ...)을 받아
# result 딕셔너리와 신규 계정 임시 비번 목록(temp_credentials)을 직접 갱신한다.
# 치명적 DB flush 오류 시 _RestoreAborted를 raise → 디스패처가 redirect로 변환.

class _RestoreAborted(Exception):
    """복원 중 치명적 DB 오류 신호. 디스패처가 catch해서 redirect 반환."""


_STATUS_KO_TO_EN = {
    '출석':         'present',
    '지각':         'late',
    '결석':         'absent',
    '조퇴':         'early_leave',
    '출석인정':     'approved_leave',
    '방과후출결인정': 'after_school',
}


def _new_restore_result():
    return {'students': 0, 'teachers': 0, 'admins': 0, 'study_rooms': 0,
            'holidays': 0, 'period_settings': 0, 'skipped': 0,
            'attendance': 0, 'applications': 0, 'study_logs': 0,
            'room_assignments': 0, 'schedules': 0, 'errors': []}


def _try_flush_or_abort(stage_name):
    try:
        db.session.flush()
    except Exception as e:
        db.session.rollback()
        flash(f'{stage_name} 복원 중 DB 오류: {e}', 'danger')
        raise _RestoreAborted()


def _restore_students_sheet(wb, result, temp_credentials):
    if '학생명단' not in wb.sheetnames:
        return
    ws = wb['학생명단']
    # 컬럼: 학번, 이름, 학년, 반, 성별, 아이디
    # (옛 백업의 7번째 비밀번호해시 컬럼은 보안상 무시 - 항상 임시 비번 새로 발급)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        sid        = str(row[0]).strip()
        name       = str(row[1]).strip()
        grade      = row[2]
        cls        = row[3]
        gender_str = row[4]
        username   = str(row[5]).strip() if row[5] else None
        if not sid or not name:
            continue
        ok, err = validate_student_id(sid)
        if not ok:
            result['errors'].append(f'학번 형식 오류 (건너뜀): {sid!r} — {err}')
            result['skipped'] += 1
            continue
        if User.query.filter_by(student_id=sid).first():
            result['skipped'] += 1
            continue
        if username and User.query.filter_by(username=username).first():
            username = sid
        username = username or sid
        gender = 'M' if str(gender_str).strip() in ('남', 'M') else 'F'
        u = User(
            username=username, name=name, role='student',
            grade=int(grade) if grade else None,
            class_num=int(cls) if cls else None,
            gender=gender, student_id=sid, is_approved=True,
        )
        temp_pw = generate_temp_password()
        u.set_password(temp_pw)
        db.session.add(u)
        temp_credentials.append(('학생', username, name, sid, temp_pw))
        result['students'] += 1
    _try_flush_or_abort('학생 계정')


def _restore_study_rooms_sheet(wb, result):
    if '자습실목록' not in wb.sheetnames:
        return
    ws = wb['자습실목록']
    # 컬럼: 자습실명, 전체정원, 남학생정원, 여학생정원, 활성화, 순서
    # (옛 백업의 7번째 QR토큰 컬럼은 보안상 무시 - 항상 새 토큰 생성, QR 재인쇄 필요)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            name        = str(row[0]).strip()
            capacity    = int(row[1]) if row[1] not in (None, '') else 0
            male_cap    = int(row[2]) if row[2] not in (None, '') else 0
            female_cap  = int(row[3]) if row[3] not in (None, '') else 0
            is_active   = str(row[4]).strip().upper() != 'N' if len(row) > 4 and row[4] else True
            order       = int(row[5]) if len(row) > 5 and row[5] not in (None, '') else 0
            if not name:
                continue
            if StudyRoom.query.filter_by(name=name).first():
                result['skipped'] += 1
                continue
            capacity   = max(0, capacity)
            male_cap   = max(0, male_cap)
            female_cap = max(0, female_cap)
            if capacity > 0 and (male_cap + female_cap) > capacity:
                result['errors'].append(
                    f'자습실 "{name}": 남/여 정원 합계({male_cap}+{female_cap})'
                    f'가 전체 정원({capacity})을 초과하여 정원값을 0으로 초기화합니다.'
                )
                male_cap = 0
                female_cap = 0
            db.session.add(StudyRoom(
                name=name, capacity=capacity,
                male_capacity=male_cap, female_capacity=female_cap,
                is_active=is_active, order=order,
                qr_token=secrets.token_hex(16),
            ))
            result['study_rooms'] += 1
        except Exception as e:
            result['errors'].append(f'자습실 행 오류: {e}')
    _try_flush_or_abort('자습실')


def _restore_attendance_sheet(wb, result, sid_to_user):
    if '출결상세' not in wb.sheetnames:
        return
    ws = wb['출결상세']
    # 컬럼: 학번, 이름, 날짜, 교시, 상태, 출석시각, 퇴실시각(선택), 조퇴사유(선택), 자습공간명, 자습시간(분)
    def _parse_dt(val):
        s = str(val).strip() if val else ''
        return datetime.strptime(s[:19], '%Y-%m-%d %H:%M:%S') if s else None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            sid      = str(row[0]).strip()
            att_date = datetime.strptime(str(row[2]).strip()[:10], '%Y-%m-%d').date()
            period   = int(row[3])
            status   = _STATUS_KO_TO_EN.get(str(row[4]).strip(), 'present')
            checked_at      = _parse_dt(row[5] if len(row) > 5 else None)
            checked_out_at  = _parse_dt(row[6] if len(row) > 6 else None)
            early_leave_note = str(row[7]).strip() if len(row) > 7 and row[7] else None
            room_name        = str(row[8]).strip() if len(row) > 8 and row[8] else None
            study_minutes    = int(row[9]) if len(row) > 9 and row[9] not in (None, '') else None

            user = sid_to_user.get(sid)
            if not user:
                continue
            if Attendance.query.filter_by(user_id=user.id, date=att_date, period=period).first():
                continue
            room_obj = StudyRoom.query.filter_by(name=room_name).first() if room_name else None
            db.session.add(Attendance(
                user_id=user.id, date=att_date, period=period,
                status=status, checked_at=checked_at,
                checked_out_at=checked_out_at,
                study_minutes=study_minutes,
                early_leave_note=early_leave_note,
                study_room_id=room_obj.id if room_obj else None,
            ))
            result['attendance'] += 1
        except Exception as e:
            result['errors'].append(f'출결 행 오류: {e}')
    _try_flush_or_abort('출결')


def _restore_applications_sheet(wb, result, sid_to_user):
    if '자습신청' not in wb.sheetnames:
        return
    ws = wb['자습신청']
    # 컬럼: 학번, 이름, 학년, 반, 날짜, 교시
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            sid      = str(row[0]).strip()
            app_date = datetime.strptime(str(row[4]).strip()[:10], '%Y-%m-%d').date()
            period   = int(row[5])
            user = sid_to_user.get(sid)
            if not user:
                continue
            if StudyApplication.query.filter_by(user_id=user.id, date=app_date, period=period).first():
                continue
            db.session.add(StudyApplication(user_id=user.id, date=app_date, period=period))
            result['applications'] += 1
        except Exception as e:
            result['errors'].append(f'신청 행 오류: {e}')
    _try_flush_or_abort('자습 신청')


def _restore_study_logs_sheet(wb, result, sid_to_user):
    if '학습기록' not in wb.sheetnames:
        return
    ws = wb['학습기록']
    # 컬럼: 학번, 이름, 학년, 반, 날짜, 과목, 학습시간(분), 메모
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            sid      = str(row[0]).strip()
            log_date = datetime.strptime(str(row[4]).strip()[:10], '%Y-%m-%d').date()
            subject  = str(row[5]).strip()
            duration = int(row[6])
            memo     = str(row[7]).strip() if row[7] else ''
            user = sid_to_user.get(sid)
            if not user or not subject:
                continue
            db.session.add(StudyLog(
                user_id=user.id, date=log_date,
                subject=subject, duration=duration, memo=memo
            ))
            result['study_logs'] += 1
        except Exception as e:
            result['errors'].append(f'학습기록 행 오류: {e}')
    _try_flush_or_abort('학습 기록')


def _restore_teachers_sheet(wb, result, temp_credentials):
    if '교사명단' not in wb.sheetnames:
        return
    ws = wb['교사명단']
    # 컬럼: 아이디, 이름, 승인여부, 담당학년(선택)
    # (옛 백업의 5번째 비밀번호해시 컬럼은 보안상 무시 - 항상 임시 비번 새로 발급)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            username        = str(row[0]).strip()
            name            = str(row[1]).strip()
            approved        = str(row[2]).strip() == '승인'
            assigned_grade  = int(row[3]) if len(row) > 3 and row[3] not in (None, '') else None
            if not username or not name:
                continue
            if User.query.filter_by(username=username).first():
                result['skipped'] += 1
                continue
            t = User(
                username=username, name=name, role='teacher',
                is_approved=approved, assigned_grade=assigned_grade,
            )
            temp_pw = generate_temp_password()
            t.set_password(temp_pw)
            db.session.add(t)
            temp_credentials.append(('교사', username, name, '', temp_pw))
            result['teachers'] += 1
        except Exception as e:
            result['errors'].append(f'교사 행 오류: {e}')
    _try_flush_or_abort('교사 계정')


def _restore_admins_sheet(wb, result, temp_credentials):
    if '관리자' not in wb.sheetnames:
        return
    ws = wb['관리자']
    # 컬럼: 아이디, 이름
    # (옛 백업의 3번째 비밀번호해시 컬럼은 보안상 무시 - 항상 임시 비번 새로 발급)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            username = str(row[0]).strip()
            name     = str(row[1]).strip() if row[1] else username
            if not username:
                continue
            if User.query.filter_by(username=username).first():
                result['skipped'] += 1
                continue
            a = User(username=username, name=name, role='admin', is_approved=True)
            temp_pw = generate_temp_password()
            a.set_password(temp_pw)
            db.session.add(a)
            temp_credentials.append(('관리자', username, name, '', temp_pw))
            result['admins'] += 1
        except Exception as e:
            result['errors'].append(f'관리자 행 오류: {e}')
    _try_flush_or_abort('관리자 계정')


def _restore_holidays_sheet(wb, result):
    if '공휴일' not in wb.sheetnames:
        return
    ws = wb['공휴일']
    # 컬럼: 날짜, 공휴일명
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            h_date = datetime.strptime(str(row[0]).strip()[:10], '%Y-%m-%d').date()
            h_name = str(row[1]).strip() if row[1] else ''
            if not h_name:
                continue
            if Holiday.query.filter_by(date=h_date).first():
                result['skipped'] += 1
                continue
            db.session.add(Holiday(date=h_date, name=h_name))
            result['holidays'] += 1
        except Exception as e:
            result['errors'].append(f'공휴일 행 오류: {e}')
    _try_flush_or_abort('공휴일')


def _restore_period_settings_sheet(wb, result):
    if '자습시간설정' not in wb.sheetnames:
        return
    ws = wb['자습시간설정']
    # 컬럼: 요일구분, 교시, 시작시각, 종료시각, 활성화
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            day_type   = str(row[0]).strip()
            period     = int(row[1])
            start_time = str(row[2]).strip()
            end_time   = str(row[3]).strip()
            is_active  = str(row[4]).strip().upper() != 'N' if len(row) > 4 and row[4] else True
            if not day_type or not start_time or not end_time:
                continue
            ok_s, err_s = validate_time_str(start_time, f'{day_type} {period}교시 시작 시각')
            if not ok_s:
                result['errors'].append(f'자습시간설정: {err_s}')
                continue
            ok_e, err_e = validate_time_str(end_time, f'{day_type} {period}교시 종료 시각')
            if not ok_e:
                result['errors'].append(f'자습시간설정: {err_e}')
                continue
            if parse_time_str(start_time) >= parse_time_str(end_time):
                result['errors'].append(f'자습시간설정 {day_type} {period}교시 시작≥종료')
                continue
            existing = StudyPeriodSetting.query.filter_by(
                day_type=day_type, period=period).first()
            if existing:
                existing.start_time = start_time
                existing.end_time   = end_time
                existing.is_active  = is_active
            else:
                db.session.add(StudyPeriodSetting(
                    day_type=day_type, period=period,
                    start_time=start_time, end_time=end_time,
                    is_active=is_active,
                ))
            result['period_settings'] += 1
        except Exception as e:
            result['errors'].append(f'자습시간설정 행 오류: {e}')
    _try_flush_or_abort('자습시간설정')


def _restore_room_assignments_sheet(wb, result):
    if '자습실배정' not in wb.sheetnames:
        return
    # 최신 sid→user 맵 (방금 복원된 학생 포함)
    sid_to_user = {u.student_id: u for u in User.query.filter_by(role='student').all()}
    room_name_map = {r.name: r for r in StudyRoom.query.all()}
    ws = wb['자습실배정']
    # 컬럼: 학번, 이름, 자습실명, 좌석번호, 위치X(%), 위치Y(%)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            sid       = str(row[0]).strip()
            room_name = str(row[2]).strip() if len(row) > 2 and row[2] else None
            seat_num  = int(row[3]) if len(row) > 3 and row[3] not in (None, '') else None
            pos_x     = float(row[4]) if len(row) > 4 and row[4] not in (None, '') else None
            pos_y     = float(row[5]) if len(row) > 5 and row[5] not in (None, '') else None

            user = sid_to_user.get(sid)
            room = room_name_map.get(room_name) if room_name else None
            if not user or not room:
                continue
            if StudentRoom.query.filter_by(user_id=user.id).first():
                result['skipped'] += 1
                continue
            db.session.add(StudentRoom(
                user_id=user.id, study_room_id=room.id,
                seat_number=seat_num, pos_x=pos_x, pos_y=pos_y,
            ))
            result['room_assignments'] += 1
        except Exception as e:
            result['errors'].append(f'자습실배정 행 오류: {e}')
    _try_flush_or_abort('자습실 배정')


def _restore_schedules_sheet(wb, result):
    if '방과후수업' not in wb.sheetnames:
        return
    sid_to_user = {u.student_id: u for u in User.query.filter_by(role='student').all()}
    # 활성 교시 화이트리스트 (없으면 검증 생략)
    valid_periods = {s.period for s in StudyPeriodSetting.query.filter_by(is_active=True).all()
                     if s.day_type in ('weekday', 'mon', 'tue', 'wed', 'thu', 'fri')}
    ws = wb['방과후수업']
    # 컬럼: 학번, 이름, 학년, 반, 요일(한글), 요일번호(0=월), 교시
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            sid     = str(row[0]).strip()
            day_i   = int(row[5]) if len(row) > 5 and row[5] not in (None, '') else None
            period  = int(row[6]) if len(row) > 6 and row[6] not in (None, '') else None
            if day_i is None or period is None:
                continue
            if not (0 <= day_i <= 4):
                result['errors'].append(f'방과후 요일 오류 (건너뜀): 학번={sid}, 요일={day_i}')
                continue
            if valid_periods and period not in valid_periods:
                result['errors'].append(f'방과후 교시 오류 (건너뜀): 학번={sid}, {period}교시는 활성 교시 아님')
                continue
            user = sid_to_user.get(sid)
            if not user:
                continue
            if Schedule.query.filter_by(user_id=user.id, day_of_week=day_i, period=period).first():
                result['skipped'] += 1
                continue
            db.session.add(Schedule(
                user_id=user.id, day_of_week=day_i, period=period,
                subject='방과후수업',
            ))
            result['schedules'] += 1
        except Exception as e:
            result['errors'].append(f'방과후수업 행 오류: {e}')
    _try_flush_or_abort('방과후수업')


def _build_restore_summary_msg(result):
    return (f'복원 완료 — 학생 {result["students"]}명 / '
            f'교사 {result["teachers"]}명 / '
            f'관리자 {result["admins"]}명 / '
            f'자습실 {result["study_rooms"]}개 / '
            f'공휴일 {result["holidays"]}건 / '
            f'자습시간설정 {result["period_settings"]}건 생성 '
            f'(중복 건너뜀 {result["skipped"]}건) / '
            f'출결 {result["attendance"]}건 / '
            f'자습신청 {result["applications"]}건 / '
            f'학습기록 {result["study_logs"]}건 / '
            f'자습실배정 {result["room_assignments"]}건 / '
            f'방과후수업 {result["schedules"]}건')


def _build_temp_credentials_xlsx(msg, result, temp_credentials):
    """신규 계정의 임시 비번을 담은 Excel 워크북을 BytesIO로 반환."""
    wb_out = openpyxl.Workbook()

    ws_sum = wb_out.active
    ws_sum.title = '복원요약'
    ws_sum.append(['항목', '값'])
    ws_sum.append(['복원 요약', msg])
    ws_sum.append(['신규 계정 수', len(temp_credentials)])
    ws_sum.append(['생성 시각', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws_sum.append(['주의', '이 파일은 모든 신규 계정의 초기 비밀번호를 평문으로 담고 있습니다. '
                          '학생·교사에게 개별 전달한 후 즉시 삭제하십시오.'])

    ws_cred = wb_out.create_sheet('임시비밀번호')
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_font = Font(bold=True, color='FFFFFF')
    ws_cred.append(['역할', '아이디', '이름', '학번', '임시비밀번호'])
    for cell in ws_cred[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
    for role, username, uname, sid, pw in temp_credentials:
        ws_cred.append([role, username, uname, sid, pw])
    ws_cred.freeze_panes = 'A2'

    if result['errors']:
        ws_err = wb_out.create_sheet('복원오류')
        ws_err.append(['행', '메시지'])
        for i, err in enumerate(result['errors'], start=1):
            ws_err.append([i, err])
        ws_err.freeze_panes = 'A2'

    buf = BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf


@admin_bp.route('/restore', methods=['GET', 'POST'])
def restore():
    """백업 Excel 파일로 DB 복원 - 디스패처 전용. 시트별 처리는 _restore_*_sheet() 헬퍼들."""
    if request.method == 'GET':
        return render_template('admin/restore.html')

    f = request.files.get('backup_file')
    if not f or not f.filename.endswith('.xlsx'):
        flash('xlsx 파일을 선택하세요.', 'danger')
        return render_template('admin/restore.html')

    options = request.form.getlist('options')
    flags = {name: (name in options) for name in (
        'students', 'teachers', 'admins', 'study_rooms',
        'holidays', 'period_settings', 'attendance',
        'applications', 'study_logs', 'room_assignments', 'schedules',
    )}

    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    except Exception as e:
        flash(f'파일을 읽을 수 없습니다: {e}', 'danger')
        return render_template('admin/restore.html')

    result = _new_restore_result()
    # 신규 발급된 임시 비밀번호 목록 (역할, 아이디, 이름, 학번, 임시비번)
    # 복원 종료 후 Excel로 다운로드된다 - 화면이나 로그에 평문 비번을 남기지 않는다.
    temp_credentials = []

    try:
        # 학생 먼저 - 출결/신청/학습기록 등이 sid_to_user를 참조하기 때문
        if flags['students']:
            _restore_students_sheet(wb, result, temp_credentials)
        sid_to_user = {u.student_id: u for u in User.query.filter_by(role='student').all()}

        # 자습실 먼저 - 출결 시트가 자습실명으로 study_room_id 조회
        if flags['study_rooms']:
            _restore_study_rooms_sheet(wb, result)

        if flags['attendance']:
            _restore_attendance_sheet(wb, result, sid_to_user)
        if flags['applications']:
            _restore_applications_sheet(wb, result, sid_to_user)
        if flags['study_logs']:
            _restore_study_logs_sheet(wb, result, sid_to_user)

        if flags['teachers']:
            _restore_teachers_sheet(wb, result, temp_credentials)
        if flags['admins']:
            _restore_admins_sheet(wb, result, temp_credentials)

        if flags['holidays']:
            _restore_holidays_sheet(wb, result)
        if flags['period_settings']:
            _restore_period_settings_sheet(wb, result)
        if flags['room_assignments']:
            _restore_room_assignments_sheet(wb, result)
        if flags['schedules']:
            _restore_schedules_sheet(wb, result)
    except _RestoreAborted:
        return redirect(url_for('admin_bp.restore'))

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'엑셀 복원 중 DB 저장 오류가 발생했습니다: {e}', 'danger')
        return redirect(url_for('admin_bp.restore'))

    log_audit('admin.excel_restore', level='warning',
              admin=current_user.username, source_filename=f.filename,
              students=result['students'], teachers=result['teachers'],
              admins=result['admins'])

    msg = _build_restore_summary_msg(result)

    # 새로 생성된 계정이 없으면 redirect+flash
    if not temp_credentials:
        flash(msg, 'success')
        if result['errors']:
            for e in result['errors'][:5]:
                flash(e, 'warning')
        return redirect(url_for('admin_bp.dashboard'))

    # 신규 계정이 있으면 임시 비번 Excel을 즉시 다운로드시킨다
    # (화면·flash·로그에 평문 비번을 남기지 않는 유일한 경로)
    buf = _build_temp_credentials_xlsx(msg, result, temp_credentials)
    filename = f'self_study_복원결과_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

@admin_bp.route('/db-backup')
def db_backup():
    """SQLite DB 파일 직접 다운로드 — 시스템 업데이트 전 완전 백업"""
    if not os.path.exists(DB_PATH):
        flash('DB 파일을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('admin_bp.dashboard'))

    buf = BytesIO()
    # sqlite3 backup API로 핫 스냅샷 생성 (쓰기 잠금 없이 안전)
    tmp_path = DB_PATH + '.bak_tmp'
    try:
        src = sqlite3.connect(DB_PATH)
        bak = sqlite3.connect(tmp_path)
        src.backup(bak)
        bak.close()
        src.close()
        with open(tmp_path, 'rb') as fh:
            buf.write(fh.read())
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    buf.seek(0)
    now_str  = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'self_study_DB_{now_str}.db'
    return send_file(buf, as_attachment=True,
                     download_name=filename,
                     mimetype='application/octet-stream')


@admin_bp.route('/db-restore', methods=['POST'])
def db_restore():
    """업로드된 .db 파일로 전체 DB 교체 — 기존 DB는 .prev 파일로 자동 보존"""
    f = request.files.get('db_file')
    if not f or not f.filename.lower().endswith('.db'):
        flash('.db 파일을 선택하세요.', 'danger')
        return redirect(url_for('admin_bp.restore'))

    # 업로드 파일을 메모리에 전부 올리지 않고 디스크로 스트리밍 저장한다.
    # (Flask는 MAX_CONTENT_LENGTH 초과 시 413으로 이미 차단)
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.db', dir=os.path.dirname(DB_PATH))
    try:
        os.close(tmp_fd)
        f.save(tmp_path)

        # 매직 바이트 검증 (앞 16바이트만 읽는다)
        with open(tmp_path, 'rb') as fh:
            magic = fh.read(16)
        if len(magic) < 16 or magic != b'SQLite format 3\x00':
            flash('유효한 SQLite DB 파일이 아닙니다.', 'danger')
            return redirect(url_for('admin_bp.restore'))

        # 필수 테이블·컬럼 검사 — models.py 실제 컬럼과 1:1 대응
        REQUIRED_COLUMNS = {
            'users': {
                'id', 'username', 'password_hash', 'name', 'role', 'is_approved',
                'session_token', 'grade', 'class_num', 'gender', 'student_id',
                'assigned_grade',
            },
            'attendance': {
                'id', 'user_id', 'date', 'period', 'status',
                'study_room_id', 'checked_at', 'checked_out_at',
                'study_minutes', 'early_leave_note',
            },
            'attendance_logs': {
                'id', 'attendance_id', 'changed_by',
                'old_status', 'new_status', 'changed_at', 'note',
            },
            'study_rooms': {
                'id', 'name', 'capacity', 'male_capacity', 'female_capacity',
                'is_active', 'order', 'qr_token',
            },
            'student_rooms': {
                'id', 'user_id', 'study_room_id',
                'seat_number', 'pos_x', 'pos_y',
            },
            'study_applications': {
                'id', 'user_id', 'date', 'period', 'applied_at',
            },
            'study_logs': {
                'id', 'user_id', 'date', 'subject', 'duration', 'memo',
            },
            'study_period_settings': {
                'id', 'day_type', 'period', 'start_time', 'end_time', 'is_active',
            },
            'holidays':  {'id', 'date', 'name'},
            'schedules': {'id', 'user_id', 'day_of_week', 'period', 'subject'},
        }
        conn = sqlite3.connect(tmp_path)
        try:
            cur = conn.cursor()
            # 테이블 존재 확인
            cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = {row[0] for row in cur.fetchall()}
            missing_tables = set(REQUIRED_COLUMNS) - tables
            if missing_tables:
                flash(f'DB 복원 실패: 필수 테이블 누락 ({", ".join(sorted(missing_tables))})', 'danger')
                return redirect(url_for('admin_bp.restore'))
            # 컬럼 존재 확인
            for tbl, req_cols in REQUIRED_COLUMNS.items():
                cur.execute(f'PRAGMA table_info({tbl})')
                existing_cols = {row[1] for row in cur.fetchall()}
                missing_cols = req_cols - existing_cols
                if missing_cols:
                    flash(
                        f'DB 복원 실패: {tbl} 테이블에 필수 컬럼 누락 '
                        f'({", ".join(sorted(missing_cols))})',
                        'danger'
                    )
                    return redirect(url_for('admin_bp.restore'))
            # DB 무결성 검사
            cur.execute('PRAGMA integrity_check')
            integrity_result = cur.fetchone()[0]
            if integrity_result != 'ok':
                flash(f'DB 복원 실패: 파일 무결성 오류 ({integrity_result})', 'danger')
                return redirect(url_for('admin_bp.restore'))
            # 외래키 참조 무결성 검사
            cur.execute('PRAGMA foreign_key_check')
            fk_errors = cur.fetchall()
            if fk_errors:
                flash(
                    f'DB 복원 실패: 외래키 참조 오류 {len(fk_errors)}건 '
                    f'(테이블: {fk_errors[0][0]})',
                    'danger'
                )
                return redirect(url_for('admin_bp.restore'))
        finally:
            conn.close()

        # 현재 DB를 .prev로 보존 — WAL 모드 안전하게 sqlite3 backup API 사용
        prev_path = DB_PATH + '.prev'
        if os.path.exists(DB_PATH):
            src_conn = sqlite3.connect(DB_PATH)
            dst_conn = sqlite3.connect(prev_path)
            try:
                src_conn.backup(dst_conn)
            finally:
                dst_conn.close()
                src_conn.close()

        # 모든 DB 커넥션 반환 후 atomic 교체
        # Windows에서 세션이 파일 잠금을 잡고 있을 수 있으므로 session.remove() 선행
        db.session.remove()
        db.engine.dispose()
        os.replace(tmp_path, DB_PATH)
        tmp_path = None  # 성공 시 삭제 생략

        # 옛 DB 호환성: SystemSetting 등 누락 테이블 생성 + 기본값 시드 + session_token 백필
        # (이 단계가 없으면 system_settings 없는 옛 DB 복원 후 모든 페이지가 500 에러)
        from app import reinitialize_after_db_change
        reinitialize_after_db_change()

    except Exception as e:
        flash(f'DB 복원 중 오류가 발생했습니다: {e}', 'danger')
        return redirect(url_for('admin_bp.restore'))
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)

    log_audit('admin.db_restore', level='warning',
              admin=current_user.username,
              source_filename=f.filename)
    flash(
        'DB가 성공적으로 복원되었습니다. '
        '이전 DB는 self_study.db.prev 파일로 보존되어 있습니다.',
        'success'
    )
    flash(
        '※ 보안상 모든 사용자(관리자 본인 포함)의 세션이 무효화됩니다. '
        '다음 페이지 이동 시 자동으로 로그인 화면으로 이동하므로, '
        '복원한 DB의 비밀번호로 새로 로그인해 주세요.',
        'warning'
    )
    return redirect(url_for('admin_bp.dashboard'))


@admin_bp.route('/users/<int:user_id>/reset-password', methods=['POST'])
def reset_password(user_id):
    """특정 학생·교사 비밀번호를 임시 비밀번호로 초기화"""
    user = db.session.get(User, user_id)
    if not user or user.role == 'admin':
        flash('처리할 수 없는 계정입니다.', 'danger')
        return redirect(url_for('admin_bp.users'))

    temp_pw = _temp_password()
    user.set_password(temp_pw)
    db.session.commit()
    log_audit('admin.pw_reset', level='warning',
              admin=current_user.username, target=user.username,
              target_role=user.role)
    # 평문 임시 비번은 이 한 번의 응답에서만 노출된다.
    # 관리자가 이 화면을 떠나거나 새로고침하면 다시 확인할 수 없다.
    # 평문 비번은 절대 로그에 남기지 않는다 — 이벤트만 기록.
    flash(
        f'[{user.name} / {user.username}] 임시 비밀번호: {temp_pw} '
        f'— 이 화면을 벗어나면 다시 확인할 수 없습니다. '
        f'지금 바로 본인에게 전달하고, 전달 후에는 스크린샷·메모를 남기지 마세요. '
        f'재확인이 필요하면 비밀번호 초기화를 한 번 더 수행해야 합니다.',
        'warning'
    )
    return redirect(url_for('admin_bp.users'))
